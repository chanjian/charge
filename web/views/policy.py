from django.shortcuts import render, redirect

from utils.link import filter_reverse
from web import models
from utils.pager import Pagination
from django import forms
from utils.bootstrap import BootStrapForm
from utils.response import BaseResponse
from django.http import JsonResponse, HttpResponse


def policy_list(request):
    queryset = models.PricePolicy.objects.all().order_by('count')
    pager = Pagination(request, queryset)
    return render(request, 'policy_list.html', {'pager': pager})


class PolicyModelForm(BootStrapForm, forms.ModelForm):
    class Meta:
        model = models.PricePolicy
        fields = ['count','discount']

    def __init__(self, *args, **kwargs):
        self.request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)

    def clean(self):
        cleaned_data = super().clean()
        count = cleaned_data.get('count')
        discount = cleaned_data.get('discount')


        if not all([count,discount]):
            return cleaned_data

        # 如果是编辑现有实例，获取creator
        creator = self.instance.creator if self.instance.pk else self.request.userdict.id

        # 检查当前管理员是否已创建相同等级
        queryset = models.PricePolicy.objects.filter(creator=creator,active=1,count=count)

        if self.instance.pk:  # 编辑时排除自身
            queryset = queryset.exclude(pk=self.instance.pk)

        if queryset.exists():
            if queryset.filter(count=count).exists():
                raise forms.ValidationError(f"您已创建过数量为'{count}'的折扣策略")

        return cleaned_data




def policy_add(request):
    if request.method == "GET":
        form = PolicyModelForm(request=request)
        return render(request, 'form4.html', {'form': form})
    form = PolicyModelForm(data=request.POST,request=request)
    if not form.is_valid():
        return render(request, 'form4.html', {'form': form})
    form.instance.creator = request.userinfo
    form.save()
    return redirect('/policy/list/')


def policy_edit(request, pk):
    instance = models.PricePolicy.objects.filter(id=pk).first()
    if request.method == "GET":
        form = PolicyModelForm(instance=instance)
        return render(request, 'form4.html', {'form': form})
    form = PolicyModelForm(data=request.POST, instance=instance)
    if not form.is_valid():
        return render(request, 'form4.html', {'form': form})
    form.save()
    return redirect(filter_reverse(request,'/policy/list/'))


def policy_delete(request):
    res = BaseResponse(status=True)
    cid = request.GET.get('cid')
    models.PricePolicy.objects.filter(id=cid).delete()
    return JsonResponse(res.dict)

def policy_upload(request):
    """批量上传【基于excel】"""
    from openpyxl import load_workbook
    #1.获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    print(file_object)

    #2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object,data_only=True)
    sheet = wb.worksheets[0]

    #第一行第一列
    # cell = sheet.cell(1,1)
    # print(cell.value)

    #3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        print(row)
        print(row[0].value,row[1].value)
        qb_num = row[0].value
        qb_discount = row[1].value

        exists = models.PricePolicy.objects.filter(count=qb_num).exists()
        if not exists:
            models.PricePolicy.objects.create(count=qb_num,discount=qb_discount)


    return redirect('policy_list')


from django.http import FileResponse, HttpResponseNotFound
import os


def policy_example_download(request):
    # 文件路径（示例：假设文件位于项目的 `media` 目录下）
    file_path = 'media/给客服设置的收货价格策略.xlsx'  # 替换为你的实际文件路径

    if os.path.exists(file_path):
        # 打开文件并返回为 FileResponse
        file = open(file_path, 'rb')
        response = FileResponse(file)

        # 设置下载文件名（可选）
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    else:
        return HttpResponseNotFound("文件不存在！")